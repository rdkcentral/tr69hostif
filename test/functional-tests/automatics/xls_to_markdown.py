#!/usr/bin/env python3
"""
Convert formatted .xls/.xlsx files to Markdown with test case grouping.
Uses openpyxl to read Excel data (resolving merged cells) and produces
grouped Markdown output. Inspired by https://github.com/microsoft/markitdown

Usage:
    python3 xls_to_markdown.py                         # convert all .xls/.xlsx in this directory
    python3 xls_to_markdown.py file1.xls file2.xlsx    # convert specific files
"""

import os
import shutil
import sys
import tempfile

from openpyxl import load_workbook


def _get_merged_cell_value(ws, row, col):
    """Return the value for a cell, resolving merged ranges."""
    cell = ws.cell(row, col)
    if cell.value is not None:
        return cell.value
    # Check if this cell is part of a merged range
    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            # Return the top-left cell's value
            return ws.cell(merged_range.min_row, merged_range.min_col).value
    return None


def _escape_md(text):
    """Escape pipe characters in markdown table cell text."""
    if text is None:
        return ""
    text = str(text)
    text = text.replace("|", "\\|")
    return text


def _read_sheet_data(ws):
    """Read all sheet data, resolving merged cells into a 2D list of strings."""
    rows = []
    for r in range(1, ws.max_row + 1):
        row_data = []
        for c in range(1, ws.max_column + 1):
            val = _get_merged_cell_value(ws, r, c)
            row_data.append(val)
        rows.append(row_data)
    return rows


def _identify_groups(rows, sno_col_idx, tc_col_idx):
    """Identify test case groups from data rows (skip header at index 0).

    Returns list of (group_sno, group_tc_name, start_idx, end_idx) where
    indices are into the rows list (1-based, skipping header).
    """
    groups = []
    current_tc = None
    current_sno = None
    for i in range(1, len(rows)):  # skip header
        sno = rows[i][sno_col_idx]
        tc = rows[i][tc_col_idx]

        # A new group starts when S.No or AUTOMATION ID has a non-empty value
        if tc is not None and tc != current_tc:
            groups.append((sno, tc, i, i))
            current_tc = tc
            current_sno = sno
        elif sno is not None and sno != current_sno:
            groups.append((sno, tc if tc else current_tc, i, i))
            current_tc = tc if tc else current_tc
            current_sno = sno
        else:
            if groups:
                # Extend the current group
                groups[-1] = (groups[-1][0], groups[-1][1], groups[-1][2], i)
    return groups


def _build_grouped_markdown(sheet_name, headers, rows, groups, sno_col_idx, tc_col_idx):
    """Build markdown string with test cases grouped under headings."""
    lines = []
    lines.append(f"## {sheet_name}\n")

    # Determine step-data column indices (everything except S.No and AUTOMATION ID)
    step_cols = [c for c in range(len(headers)) if c != sno_col_idx and c != tc_col_idx]
    step_headers = [headers[c] for c in step_cols]

    # Build the step table header
    header_line = "| " + " | ".join(_escape_md(h) for h in step_headers) + " |"
    sep_line = "| " + " | ".join("---" for _ in step_headers) + " |"

    for sno, tc_name, start, end in groups:
        # Group heading: S.No and test case name
        sno_str = int(sno) if isinstance(sno, float) and sno == int(sno) else sno
        tc_display = str(tc_name).replace("_", "\\_") if tc_name else "Unknown"
        lines.append(f"### {sno_str}. {tc_display}\n")

        # Step table
        lines.append(header_line)
        lines.append(sep_line)
        for r in range(start, end + 1):
            cells = [_escape_md(rows[r][c]) for c in step_cols]
            lines.append("| " + " | ".join(cells) + " |")
        lines.append("")  # blank line after table

    return "\n".join(lines)


def convert_to_markdown(filepath, output_path=None):
    """Convert a formatted Excel file to grouped Markdown."""
    # openpyxl needs .xlsx extension — handle .xls files
    _, ext = os.path.splitext(filepath)
    tmp_path = None
    if ext.lower() == ".xls":
        tmp_path = tempfile.mktemp(suffix=".xlsx")
        shutil.copy2(filepath, tmp_path)
        load_path = tmp_path
    else:
        load_path = filepath

    try:
        wb = load_workbook(load_path, data_only=True)
    finally:
        pass  # cleanup happens below

    if output_path is None:
        base, _ = os.path.splitext(filepath)
        output_path = base + ".md"

    md_parts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row < 2:
            continue

        rows = _read_sheet_data(ws)
        headers = [str(h) if h else "" for h in rows[0]]

        # Find S.No and AUTOMATION ID columns
        sno_col = None
        tc_col = None
        for i, h in enumerate(headers):
            hl = h.strip().lower()
            if hl in ("s.no", "s.no.", "sno", "serial", "sl.no"):
                sno_col = i
            elif hl in ("automation id", "automation_id", "test case", "testcase"):
                tc_col = i

        if sno_col is None or tc_col is None:
            # Fallback: assume col 0 = S.No, col 1 = AUTOMATION ID
            sno_col = sno_col if sno_col is not None else 0
            tc_col = tc_col if tc_col is not None else 1

        groups = _identify_groups(rows, sno_col, tc_col)
        md_parts.append(_build_grouped_markdown(sheet_name, headers, rows, groups, sno_col, tc_col))

    wb.close()
    if tmp_path and os.path.exists(tmp_path):
        os.remove(tmp_path)

    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n".join(md_parts))

    print(f"Converted: {filepath} -> {output_path}")
    return output_path


def main():
    if len(sys.argv) > 1:
        files = sys.argv[1:]
    else:
        # Default: process all .xls/.xlsx files in the script's directory
        script_dir = os.path.dirname(os.path.abspath(__file__))
        files = [
            os.path.join(script_dir, f)
            for f in sorted(os.listdir(script_dir))
            if f.endswith((".xls", ".xlsx"))
            and not f.startswith("~")
            and not f.endswith(".bak")
        ]

    if not files:
        print("No .xls/.xlsx files found to convert.")
        sys.exit(1)

    for filepath in files:
        filepath = os.path.abspath(filepath)
        if not os.path.isfile(filepath):
            print(f"Skipping (not found): {filepath}")
            continue
        convert_to_markdown(filepath)

    print("Done.")


if __name__ == "__main__":
    main()
