#!/usr/bin/env python3
"""
Format .xls files in the automatics directory:
  1. Merge consecutive rows in column A (AUTOMATION ID) where test case name is the same.
  2. Insert a new serial number column on the left.
Overwrites the original file in place.
"""

import os
import sys
import shutil
import tempfile
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill


def format_workbook(filepath):
    """Format all sheets in the given .xls/.xlsx workbook."""
    # openpyxl requires .xlsx extension — copy to temp file if needed
    _, ext = os.path.splitext(filepath)
    tmp_path = None
    if ext.lower() == ".xls":
        tmp_path = tempfile.mktemp(suffix=".xlsx")
        shutil.copy2(filepath, tmp_path)
        wb = load_workbook(tmp_path)
    else:
        wb = load_workbook(filepath)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        _format_sheet(ws)

    # Save back — write to temp then overwrite original
    out_tmp = tempfile.mktemp(suffix=".xlsx")
    wb.save(out_tmp)
    wb.close()
    shutil.move(out_tmp, filepath)

    # Clean up temp input copy
    if tmp_path and os.path.exists(tmp_path):
        os.remove(tmp_path)

    print(f"Formatted: {filepath}")


def _format_sheet(ws):
    """Apply merging and serial number insertion to a single sheet."""
    if ws.max_row < 2:
        return  # nothing to process

    # --- Step 1: Identify consecutive groups in column 1 (AUTOMATION ID) ---
    groups = []  # list of (value, start_row, end_row)  — 1-indexed, data starts at row 2
    current_val = None
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row, 1).value
        if val == current_val:
            groups[-1] = (current_val, groups[-1][1], row)
        else:
            groups.append((val, row, row))
            current_val = val

    # --- Step 2: Insert a new column A for serial numbers ---
    ws.insert_cols(1)

    # Header for the new column
    header_cell = ws.cell(1, 1)
    header_cell.value = "S.No"
    header_cell.font = Font(bold=True)
    header_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Style the existing header row (shifted right by 1)
    for col in range(2, ws.max_column + 1):
        cell = ws.cell(1, col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # --- Step 3: Merge cells and add serial numbers ---
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for idx, (val, start_row, end_row) in enumerate(groups, start=1):
        # Serial number column (column 1) — merge and set value
        if start_row < end_row:
            ws.merge_cells(
                start_row=start_row, start_column=1,
                end_row=end_row, end_column=1,
            )
        sno_cell = ws.cell(start_row, 1)
        sno_cell.value = idx
        sno_cell.alignment = Alignment(horizontal="center", vertical="center")
        sno_cell.border = thin_border

        # AUTOMATION ID column (now column 2) — merge consecutive identical values
        if start_row < end_row:
            ws.merge_cells(
                start_row=start_row, start_column=2,
                end_row=end_row, end_column=2,
            )
        tc_cell = ws.cell(start_row, 2)
        tc_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        tc_cell.border = thin_border

    # --- Step 4: Apply borders to all data cells ---
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row, col)
            cell.border = thin_border

    # Auto-adjust column widths
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = ws.cell(1, col_idx).column_letter
        for row_idx in range(1, ws.max_row + 1):
            val = ws.cell(row_idx, col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)


def main():
    if len(sys.argv) > 1:
        files = sys.argv[1:]
    else:
        # Default: process all .xls files in the script's directory
        script_dir = os.path.dirname(os.path.abspath(__file__))
        files = [
            os.path.join(script_dir, f)
            for f in os.listdir(script_dir)
            if f.endswith((".xls", ".xlsx")) and not f.startswith("~")
        ]

    if not files:
        print("No .xls/.xlsx files found to process.")
        sys.exit(1)

    for filepath in files:
        filepath = os.path.abspath(filepath)
        if not os.path.isfile(filepath):
            print(f"Skipping (not found): {filepath}")
            continue
        format_workbook(filepath)

    print("Done.")


if __name__ == "__main__":
    main()
